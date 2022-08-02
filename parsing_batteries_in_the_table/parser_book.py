from openpyxl import Workbook, load_workbook

class BookExcel:
    def __init__(self) -> None:
        self.wb = Workbook()
        self.wb.save('information.xlsx')

    def __load_book__ (self) -> None:
        self.wb = load_workbook ('information.xlsx')
        self.ws = self.wb.active

    def __save_book__ (self) -> None:
        self.wb.save('information.xlsx')
        
    def __name_line__ (self) -> None:
        self.__load_book__()
        self.ws["A1"] = "Название товара"
        self.ws["B1"] = "Полная цена"
        self.ws["C1"] = "Скидочная цена"
        self.ws["D1"] = "Артикул"
        self.ws["E1"] = "Стартовый ток"
        self.ws["F1"] = "Ёмкость"
        self.ws["G1"] = "Номинальное напряжение"
        self.ws["H1"] = "Клеммы"
        self.ws["I1"] = "Полярность"
        self.ws["J1"] = "Длина"
        self.ws["K1"] = "Ширина"
        self.ws["L1"] = "Высота"
        self.ws["M1"] = "Вес"
        self.ws["N1"] = "Изготовитель"
        self.ws["O1"] = "Страна производства"
        self.__save_book__()

    def __name_price__ (self,name:str,price_full:str,price_sell:str, iteration_all) ->None:
        self.__load_book__()
        iteration_all +=1
        self.ws[f"A{iteration_all}"] = name
        self.ws[f"B{iteration_all}"] = price_full
        self.ws[f"C{iteration_all}"] = price_sell
        self.__save_book__()

    def __with_book__ (self,parameter:str,iteration_all, iteration_too) -> None:
       self.__load_book__()
       iteration_all +=1
       iteration_too +=1 
       match iteration_too:
            case 2:
                self.ws[f"D{iteration_all}"] = parameter.replace("Артикул: ", "")
            case 3:
                self.ws[f"E{iteration_all}"] = parameter.replace("Стартовый ток: ", "")
            case 4:
                self.ws[f"F{iteration_all}"] = parameter.replace("Емкость: ", "")
            case 5:
                self.ws[f"G{iteration_all}"] = parameter.replace("Номинальное напряжение: ", "")
            case 6:
                self.ws[f"H{iteration_all}"] = parameter.replace("Клеммы: ", "")
            case 7:
                self.ws[f"I{iteration_all}"] = parameter.replace("Полярность: ", "")
            case 8:
                self.ws[f"J{iteration_all}"] = parameter.replace("Длина: ", "")
            case 9:
                self.ws[f"K{iteration_all}"] = parameter.replace("Ширина: ", "")
            case 10:
                self.ws[f"L{iteration_all}"] = parameter.replace("Высота: ", "")
            case 11:
                self.ws[f"M{iteration_all}"] = parameter.replace("Вес: ", "")
            case 12:
                self.ws[f"N{iteration_all}"] = parameter.replace("Изготовитель: ", "")
            case 13:
                self.ws[f"O{iteration_all}"] = parameter.replace("Страна производства: ", "")